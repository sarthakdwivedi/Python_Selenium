
#imports
import base64
import time
import openpyxl


from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

import config

#declaration of variable
book = None
sheet = None

#Function to handle open browser supporting chrome,chrome headless and firefox. can be extended to handle other browsers as well
def openBrowser():
    driver = None;
    browserType = config.browser.lower();
    if browserType == "chrome":

        #checking if headless
        if config.headless:
            chrome_options = Options();
            chrome_options.add_argument("--headless")
            driver = webdriver.Chrome(executable_path=config.chrome_location, options=chrome_options)
        else:
            driver = webdriver.Chrome(executable_path=config.chrome_location)
    elif browserType == "firefox":
        driver = webdriver.Firefox(config.firefox_location)
    else:
        print(str(browserType) + " is not supported")
    driver.maximize_window()
    return driver;

#Function to take full page screenshot
def takeFullScreenShot(driver, fileName):
    try:
        S = lambda X: driver.execute_script('return document.body.parentNode.scroll' + X)
        driver.set_window_size(S('Width'), S(
            'Height'))  # May need manual adjustment
        driver.find_element_by_tag_name('body').screenshot(fileName)

    except NoSuchElementException:
        print("Element Not found")

#function to download image
def downloadImage(data, fileName):
    imgdata = base64.b64decode(data)
    # filename = 'some_image.jpg'  # I assume you have a way of picking unique filenames
    with open(fileName, 'wb') as f:
        f.write(imgdata)

#function to get locator type
def getType(locatorType):
    locatorType = locatorType.lower()
    if locatorType == "id":
        return By.ID
    elif locatorType == "xpath":
        return By.XPATH
    elif locatorType == "name":
        return By.NAME
    elif locatorType == "cssselector":
        return By.CSS_SELECTOR
    elif locatorType == "linktext":
        return By.LINK_TEXT
    elif locatorType == "partiallinktext":
        return By.PARTIAL_LINK_TEXT
    elif locatorType == "tag":
        return By.TAG_NAME
    elif locatorType == "classname":
        return By.CLASS_NAME
    else:
        print("LocatorType " + locatorType + " not supported")
    return False

#function to get Element based on locator type and value
def getElement(driver, locatorType, locatorValue):
    element = None
    try:
        byType = getType(locatorType)
        element = driver.find_element(byType, locatorValue)
        #print("Element found")

    except:
        print("Element not found")
        raise NoSuchElementException
    finally:
        return element

#Function to find all the elements

def getElements(driver, locatorType, locatorValue):
    elements = []
    try:
        byType = getType(locatorType)
        elements = driver.find_elements(byType, locatorValue)
        #print("Element found")
    except:
        print("Element not found")
    return elements



#function to take screenshot with provided name and location
def takeScreenshot(driver, location,fileName=time.strftime("%Y%m%d-%H%M%S") + ".png"):
    fileName=location+"/"+fileName;
    try:
        driver.save_screenshot(fileName)
        #print("Screenshot saved successfully")
    except:
        print("Screenshot not saved!")


#function to get Test data from input excel file and return a dictionary
def getDataForTestCaseFromExcel(sheetName,testName):
    # Loading workbook
    book = openpyxl.load_workbook(config.excelFilePath)
    sheet = book.get_sheet_by_name(sheetName)
    row = getRowForTestCase(sheet, testName)
    data=dict()
    index=1;
    if row is not None:
        #iterating columns
        for col in sheet.iter_cols(min_row=1,max_row=1,min_col=2,max_col=len(sheet[1])):
            for cell in col:
                #storing cell and value
                data[str(cell.value).lower()]=row[index].value
                index+=1;
    return data

#getting row for the test case
def getRowForTestCase(sheet, testName):

    for row in sheet.iter_rows(min_row=1,max_row=len(sheet['A'])):
        for cell in row:
            if cell.value == testName:
                return row
    else:
        print("test case not present in the sheet")
        return None;

#function to write text in element
def writeText(element,text):
    element.clear()
    element.send_keys(text)

#function to write text in element found by locator details
def writeText(driver, locatorType, locatorValue,text):
    element=getElement(driver,locatorType,locatorValue)
    element.clear()
    element.send_keys(text)

#function to wait till element is visible
def waitForElementToBeVisible(driver, locatorType, locatorValue):
    try:
        WebDriverWait(driver, 10).until(
        expected_conditions.presence_of_element_located((getType(locatorType), locatorValue)))
    except TimeoutException:
        print("Timeout while waiting for element:"+str(locatorValue))
        raise TimeoutException

#Function to wait till element is clickable
def waitForElementToBeClickable(driver, locatorType, locatorValue):
    try:
        #element=getElement(driver,locatorType,locatorValue)
        WebDriverWait(driver, int(config.Timeout)).until(
            expected_conditions.element_to_be_clickable((getType(locatorType), locatorValue)))
        #element.click();
    except TimeoutException:
        print(str(locatorValue)+" is not clickable")
        raise TimeoutException

#function to click element
def click(driver,locatorType, locatorValue):
    waitForElementToBeClickable(driver, locatorType, locatorValue)
    element=getElement(driver, locatorType, locatorValue)
    element.click();

#function to read text from element
def readText(driver,locatorType, locatorValue):
    element=getElement(driver, locatorType, locatorValue)
    #print(element)
    #print(element.text)
    return str(element.text)
